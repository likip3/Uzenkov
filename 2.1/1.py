import csv
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

currencyTransfer = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class Vacancy:
    def __init__(self, vacancies):
        self.name = vacancies[name_list[0]]
        self.publicationYear = int(vacancies[name_list[5]][:4])
        self.averageSalary = math.floor((float(vacancies[name_list[1]]) + float(vacancies[name_list[2]])) / 2) * \
                             currencyTransfer[vacancies[name_list[3]]]
        self.areaName = vacancies[name_list[4]]
        self.publicationYear = int(vacancies[name_list[5]][:4])


class DataSet:
    def __init__(self, filename, vacancyName):
        self.filename = filename
        self.vacancyName = vacancyName

    def csv_reader(self):
        # region Val
        count = 0
        salaryByName = {}
        numberByName = {}
        header = []
        salary = {}
        city = {}
        number = {}
        vacancyNum = {}
        dynamicsSalByYears = {}
        dynamicsSalByYearsForVac = {}
        dynamicsSalByCityDescOrder = {}
        dynamicsRateByCity = {}
        # endregion
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            for index, row in enumerate(reader):
                if index == 0:
                    headersLength = len(row)
                    header = row
                elif '' not in row and len(row) == headersLength:
                    vacancies = Vacancy(dict(zip(header, row)))
                    if vacancies.publicationYear in salary:
                        salary[vacancies.publicationYear].append(vacancies.averageSalary)
                    else:
                        salary[vacancies.publicationYear] = [vacancies.averageSalary]

                    if vacancies.publicationYear in vacancyNum:
                        vacancyNum[vacancies.publicationYear] += 1
                    else:
                        vacancyNum[vacancies.publicationYear] = 1

                    if vacancies.areaName in number:
                        number[vacancies.areaName] += 1
                    else:
                        number[vacancies.areaName] = 1

                    if vacancies.areaName in city:
                        city[vacancies.areaName].append(vacancies.averageSalary)
                    else:
                        city[vacancies.areaName] = [vacancies.averageSalary]

                    if vacancies.name.find(self.vacancyName) != -1:
                        if vacancies.publicationYear in salaryByName:
                            salaryByName[vacancies.publicationYear].append(vacancies.averageSalary)
                        else:
                            salaryByName[vacancies.publicationYear] = [vacancies.averageSalary]

                        if vacancies.publicationYear in numberByName:
                            numberByName[vacancies.publicationYear] += 1
                        else:
                            numberByName[vacancies.publicationYear] = 1
                    count += 1

        if not salaryByName:
            numberByName = dict([(k, 0) for k, v in vacancyNum.copy().items()])
            salaryByName = dict([(k, []) for k, v in salary.copy().items()])
        for year, salaries in salaryByName.items():
            dynamicsSalByYearsForVac[year] = 0 if len(salaries) == 0 else int(sum(salaries) / len(salaries))
        for year, salaries in salary.items():
            dynamicsSalByYears[year] = int(sum(salaries) / len(salaries))
        for year, salaries in number.items():
            dynamicsRateByCity[year] = round(salaries / count, 4)
        for year, salaries in city.items():
            dynamicsSalByCityDescOrder[year] = int(sum(salaries) / len(salaries))

        dynamicsRateByCity = list(
            filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in dynamicsRateByCity.items()]))
        dynamicsRateByCity.sort(key=lambda a: a[-1], reverse=True)
        dynamicsSalByCityDescOrder = list(
            filter(lambda a: a[0] in list(dict(dynamicsRateByCity).keys()),
                   [(key, value) for key, value in dynamicsSalByCityDescOrder.items()]))
        dynamicsSalByCityDescOrder.sort(key=lambda a: a[-1], reverse=True)

        print('Динамика уровня зарплат по годам: {0}'.format(dynamicsSalByYears))
        print('Динамика количества вакансий по годам: {0}'.format(vacancyNum))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(dynamicsSalByYearsForVac))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(numberByName))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(dict(dynamicsSalByCityDescOrder[:10])))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(dict(dynamicsRateByCity.copy()[:10])))

        return dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, \
               dict(dynamicsSalByCityDescOrder[:10]), dict(dynamicsRateByCity.copy()[:10])


class InputStart:
    def __init__(self):
        self.filename = input('Введите название файла: ')
        self.vacancyName = input('Введите название профессии: ')

        dataset = DataSet(self.filename, self.vacancyName)
        dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10 = dataset.csv_reader()

        reportFin = Report(self.vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                           dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10)
        reportFin.generate_excel()


class Report:
    def __init__(self, vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                 dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10):
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.SalByYears, self.vacancyNum, self.SalByYearsForVac, self.numberByName, self.SalByCityDescOrder, self.RateByCity \
            = dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10

    def generate_excel(self):
        widths = []

        workSheetYear = self.workbook.active

        workSheetYear.title = 'Статистика по годам'
        workSheetYear.append(
            ['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
             'Количество вакансий - ' + self.vacancyName])
        for year in self.SalByYears.keys():
            workSheetYear.append(
                [year, self.SalByYears[year], self.SalByYearsForVac[year], self.vacancyNum[year],
                 self.numberByName[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        for row in data:
            for i, cell in enumerate(row):
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]

        for i, column_width in enumerate(widths, 1):
            workSheetYear.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.SalByCityDescOrder.items(), self.RateByCity.items()):
            data.append([city1, value1, '', city2, value2])
        workSheetCity = self.workbook.create_sheet('Статистика по городам')
        for row in data:
            workSheetCity.append(row)

        widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]

        for i, column_width in enumerate(widths, 1):
            workSheetCity.column_dimensions[get_column_letter(i)].width = column_width + 2

        bold = Font(bold=True)

        for index, _ in enumerate(self.SalByCityDescOrder):
            workSheetCity['E' + str(index + 2)].number_format = '0.00%'

        for column in 'ABCDE':
            workSheetYear[column + '1'].font = bold
            workSheetCity[column + '1'].font = bold

        slim = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for column in 'ABDE':
                workSheetCity[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.SalByYears[1] = 1
        for row, _ in enumerate(self.SalByYears):
            for column in 'ABCDE':
                workSheetYear[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')



InputStart()
