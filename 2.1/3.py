import csv
import math
import matplotlib.pyplot as plt
import numpy as np
import pathlib
import pdfkit
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from jinja2 import Environment, FileSystemLoader

name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

currencyTransfer = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

listHeaders = ['Динамика уровня зарплат по годам: ', 'Динамика количества вакансий по годам: ',
               'Динамика уровня зарплат по годам для выбранной профессии: ',
               'Динамика количества вакансий по годам для выбранной профессии: ',
               'Уровень зарплат по городам (в порядке убывания): ', 'Доля вакансий по городам (в порядке убывания): ']


class Vacancy:
    def __init__(self, vacancies):
        self.name = vacancies[name_list[0]]
        self.averageSalary = math.floor((float(vacancies[name_list[1]]) + float(vacancies[name_list[2]])) / 2) * \
                             currencyTransfer[vacancies[name_list[3]]]
        self.areaName = vacancies[name_list[4]]
        self.publicationYear = int(vacancies[name_list[5]][:4])


class DataSet:
    def __init__(self, filename, vacancyName):
        self.filename = filename
        self.vacancyName = vacancyName

    def csv_reader(self):
        # region Val
        count = 0
        salaryByName = {}
        numberByName = {}
        header = []
        salary = {}
        city = {}
        number = {}
        vacancyNum = {}
        dynamicsSalByYears = {}
        dynamicsSalByYearsForVac = {}
        dynamicsSalByCityDescOrder = {}
        dynamicsRateByCity = {}
        # endregion
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            for index, row in enumerate(reader):
                if index == 0:
                    headersLength = len(row)
                    header = row
                elif '' not in row and len(row) == headersLength:
                    vacancies = Vacancy(dict(zip(header, row)))
                    if vacancies.publicationYear in salary:
                        salary[vacancies.publicationYear].append(vacancies.averageSalary)
                    else:
                        salary[vacancies.publicationYear] = [vacancies.averageSalary]

                    if vacancies.publicationYear in vacancyNum:
                        vacancyNum[vacancies.publicationYear] += 1
                    else:
                        vacancyNum[vacancies.publicationYear] = 1

                    if vacancies.areaName in number:
                        number[vacancies.areaName] += 1
                    else:
                        number[vacancies.areaName] = 1

                    if vacancies.areaName in city:
                        city[vacancies.areaName].append(vacancies.averageSalary)
                    else:
                        city[vacancies.areaName] = [vacancies.averageSalary]

                    if vacancies.name.find(self.vacancyName) != -1:
                        if vacancies.publicationYear in salaryByName:
                            salaryByName[vacancies.publicationYear].append(vacancies.averageSalary)
                        else:
                            salaryByName[vacancies.publicationYear] = [vacancies.averageSalary]

                        if vacancies.publicationYear in numberByName:
                            numberByName[vacancies.publicationYear] += 1
                        else:
                            numberByName[vacancies.publicationYear] = 1
                    count += 1

        if not salaryByName:
            numberByName = dict([(k, 0) for k, v in vacancyNum.copy().items()])
            salaryByName = dict([(k, []) for k, v in salary.copy().items()])
        for year, salaries in salaryByName.items():
            dynamicsSalByYearsForVac[year] = 0 if len(salaries) == 0 else int(sum(salaries) / len(salaries))
        for year, salaries in salary.items():
            dynamicsSalByYears[year] = int(sum(salaries) / len(salaries))
        for year, salaries in number.items():
            dynamicsRateByCity[year] = round(salaries / count, 4)
        for year, salaries in city.items():
            dynamicsSalByCityDescOrder[year] = int(sum(salaries) / len(salaries))

        dynamicsRateByCity = list(
            filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in dynamicsRateByCity.items()]))
        dynamicsRateByCity.sort(key=lambda a: a[-1], reverse=True)
        dynamicsSalByCityDescOrder = list(
            filter(lambda a: a[0] in list(dict(dynamicsRateByCity).keys()),
                   [(key, value) for key, value in dynamicsSalByCityDescOrder.items()]))
        dynamicsSalByCityDescOrder.sort(key=lambda a: a[-1], reverse=True)

        print('Динамика уровня зарплат по годам: {0}'.format(dynamicsSalByYears))
        print('Динамика количества вакансий по годам: {0}'.format(vacancyNum))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(dynamicsSalByYearsForVac))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(numberByName))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(dict(dynamicsSalByCityDescOrder[:10])))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(dict(dynamicsRateByCity.copy()[:10])))

        return dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, \
               dict(dynamicsSalByCityDescOrder[:10]), dict(dynamicsRateByCity.copy()[:10])

    def csv_reader_raw(self):
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            headerLength = len(header)
            for row in reader:
                if '' not in row and len(row) == headerLength:
                    yield dict(zip(header, row))

    @staticmethod
    def increment(dictIN, k, c):
        if k in dictIN:
            dictIN[k] += c
        else:
            dictIN[k] = c

    @staticmethod
    def mean(dictAV):
        newDict = {}
        for k, v in dictAV.items():
            newDict[k] = int(sum(v) / len(v))
        return newDict

    def getDynamics(self):
        salary = {}
        salaryByName = {}
        city = {}
        count = 0

        for vacancyDictionary in self.csv_reader_raw():
            vacancy = Vacancy(vacancyDictionary)
            self.increment(salary, vacancy.publicationYear, [vacancy.averageSalary])
            if vacancy.name.find(self.vacancyName) != -1:
                self.increment(salaryByName, vacancy.publicationYear, [vacancy.averageSalary])
            self.increment(city, vacancy.areaName, [vacancy.averageSalary])
            count += 1

        numberByName = dict([(k, len(v)) for k, v in salaryByName.items()])
        vacancyNum = dict([(k, len(v)) for k, v in salary.items()])

        if not salaryByName:
            numberByName = dict([(k, 0) for k, v in vacancyNum.items()])
            salaryByName = dict([(k, [0]) for k, v in salary.items()])

        dynamicsSalByYears = self.mean(salary)
        dynamicsSalByYearsForVac = self.mean(salaryByName)
        dynamicsSalByCityDescOrder = self.mean(city)       # not mega abobus
        numVacByYearForVac = {}

        for y, s in city.items():
            numVacByYearForVac[y] = round(len(s) / count, 4)
        numVacByYearForVac = list(filter(lambda x: x[-1] >= 0.01, [(k, v) for k, v in numVacByYearForVac.items()]))
        numVacByYearForVac.sort(key=lambda x: x[-1], reverse=True)
        dynamicsRateByCity = dict(numVacByYearForVac.copy()[:10])
        numVacByYearForVac = dict(numVacByYearForVac)
        dynamicsSalByCityDescOrder = list(filter(lambda x: x[0] in list(numVacByYearForVac.keys()),
                                                 [(k, v) for k, v in dynamicsSalByCityDescOrder.items()]))
        dynamicsSalByCityDescOrder.sort(key=lambda x: x[-1], reverse=True)
        dynamicsSalByCityDescOrder = dict(dynamicsSalByCityDescOrder[:10])

        return dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrder, dynamicsRateByCity

    @staticmethod
    def print_statistic(salByYear, numByYear, salByYearForVac, numVacByYearForVac, dynamicsSalByCityDescOrder,
                        dynamicsRateByCity):
        listDynamics = [salByYear, numByYear, salByYearForVac, numVacByYearForVac, dynamicsSalByCityDescOrder,
                        dynamicsRateByCity]
        for i in range(len(listHeaders)):
            print(listHeaders[i] + '{0}'.format(listDynamics[i]))


class InputStart:
    def __init__(self):
        self.filename = input('Введите название файла: ')
        self.vacancyName = input('Введите название профессии: ')

        dataset = DataSet(self.filename, self.vacancyName)
        dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrder, dynamicsRateByCity = dataset.getDynamics()
        dataset.print_statistic(dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                                dynamicsSalByCityDescOrder, dynamicsRateByCity)

        self.report = Report(self.vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                             dynamicsSalByCityDescOrder, dynamicsRateByCity)
        # report.generate_excel() old
        self.report.generate_image()
        self.report.generateExelFromSheets()
        self.report.generatePDF()


class Report:
    def __init__(self, vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                 dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10):
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.SalByYears, self.vacancyNum, self.SalByYearsForVac, self.numberByName, self.SalByCityDescOrder, self.RateByCity \
            = dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10

    def generate_excel(self):
        widths = []

        workSheetYear = self.workbook.active

        workSheetYear.title = 'Статистика по годам'
        workSheetYear.append(
            ['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
             'Количество вакансий - ' + self.vacancyName])
        for year in self.SalByYears.keys():
            workSheetYear.append(
                [year, self.SalByYears[year], self.SalByYearsForVac[year], self.vacancyNum[year],
                 self.numberByName[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        for row in data:
            for i, cell in enumerate(row):
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]

        for i, column_width in enumerate(widths, 1):
            workSheetYear.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.SalByCityDescOrder.items(), self.RateByCity.items()):
            data.append([city1, value1, '', city2, value2])
        workSheetCity = self.workbook.create_sheet('Статистика по городам')
        for row in data:
            workSheetCity.append(row)

        widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]

        for i, column_width in enumerate(widths, 1):
            workSheetCity.column_dimensions[get_column_letter(i)].width = column_width + 2

        bold = Font(bold=True)

        for index, _ in enumerate(self.SalByCityDescOrder):
            workSheetCity['E' + str(index + 2)].number_format = '0.00%'

        for column in 'ABCDE':
            workSheetYear[column + '1'].font = bold
            workSheetCity[column + '1'].font = bold

        slim = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for column in 'ABDE':
                workSheetCity[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.SalByYears[1] = 1
        for row, _ in enumerate(self.SalByYears):
            for column in 'ABCDE':
                workSheetYear[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')

    def getYearSheet(self):

        work_sheet1 = self.workbook.active
        work_sheet1.title = 'Статистика по годам'
        work_sheet1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
                            'Количество вакансий - ' + self.vacancyName])
        for year in self.SalByYears.keys():
            work_sheet1.append(
                [year, self.SalByYears[year], self.SalByYearsForVac[year], self.vacancyNum[year],
                 self.numberByName[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            work_sheet1.column_dimensions[get_column_letter(i)].width = column_width + 2
        return work_sheet1

    def getCitySheet(self):

        new_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (city1, value1), (city2, value2) in zip(self.SalByCityDescOrder.items(), self.RateByCity.items()):
            new_data.append([city1, value1, '', city2, value2])
        work_sheet2 = self.workbook.create_sheet('Статистика по городам')
        for r in new_data:
            work_sheet2.append(r)

        column_widths = []
        for r in new_data:
            for i, cell in enumerate(r):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            work_sheet2.column_dimensions[get_column_letter(i)].width = column_width + 2

        return work_sheet2, len(new_data)

    def generateExelFromSheets(self):
        yearSheet = self.getYearSheet()
        citySheet, len_new_data = self.getCitySheet()
        bold = Font(bold=True)
        for c in 'ABCDE':
            yearSheet[c + '1'].font = bold
            citySheet[c + '1'].font = bold

        for index, _ in enumerate(self.SalByCityDescOrder):
            citySheet['E' + str(index + 2)].number_format = '0.00%'

        slim = Side(border_style='thin', color='00000000')

        for row in range(len_new_data):
            for column in 'ABDE':
                citySheet[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.SalByYears[1] = 1
        for row, _ in enumerate(self.SalByYears):
            for column in 'ABCDE':
                yearSheet[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')

    def generate_image(self):

        x = np.arange(len(self.SalByYears.keys()))
        width = 0.35
        areas = []

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(ncols=2, nrows=2)

        ax1.bar(x - width / 2, self.SalByYears.values(), width, label='средняя з/п')
        ax1.bar(x + width / 2, self.SalByYearsForVac.values(), width, label='з/п {0}'.format(self.vacancyName))
        ax1.set_title('Уровень зарплат по годам')
        ax1.set_xticks(x, self.SalByYears.keys(), rotation=90)
        ax1.legend(fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, self.vacancyNum.values(), width, label='количество вакансий')
        ax2.bar(x + width / 2, self.numberByName.values(), width,
                label='количество вакансий\n{0}'.format(self.vacancyName))
        ax2.set_xticks(x, self.vacancyNum.keys(), rotation=90)
        ax2.set_title('Количество вакансий по годам')
        ax2.legend(fontsize=8)
        ax2.grid(axis='y')
        fig.tight_layout()

        for area in self.SalByCityDescOrder.keys():
            areas.append(str(area).replace(' ', '\n').replace('-', '-\n'))
        y_pos = np.arange(len(areas))
        performance = self.SalByCityDescOrder.values()
        error = np.random.rand(len(areas))
        ax3.barh(y_pos, performance, xerr=error, align='center')
        ax3.set_title('Уровень зарплат по городам')
        ax3.set_yticks(y_pos, labels=areas, size=6)
        ax3.invert_yaxis()
        ax3.grid(axis='x')

        val = list(self.RateByCity.values()) + [1 - sum(list(self.RateByCity.values()))]
        k = list(self.RateByCity.keys()) + ['Другие']
        colors = plt.get_cmap('Paired_r')(np.linspace(0.2, 0.7, 11))
        ax4.pie(val, labels=k, startangle=170, colors=colors, textprops={'fontsize': 6})
        ax4.set_title('Доля вакансий по городам')
        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def generatePDF(self):

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        dynamics = []
        for year in self.vacancyNum.keys():
            dynamics.append(
                [year, self.SalByYears[year], self.vacancyNum[year], self.SalByYearsForVac[year],
                 self.numberByName[year]])

        for key in self.RateByCity:
            self.RateByCity[key] = round(self.RateByCity[key] * 100, 2)

        pdf_template = template.render({'vacancyName': self.vacancyName,
                                        'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                                        'workbook': self.workbook})

        config = pdfkit.configuration(
            wkhtmltopdf=r'C:\Users\likip\Downloads\wkhtmltox-0.12.6-1.mxe-cross-win64\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


InputStart()
